import csv
import time
import io
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait  # available since 2.4.0
from selenium.webdriver.support import expected_conditions as EC  # available since 2.26.0
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

path = ""
# insert the path to your workbook in the double quotes Ex. C:/FantasyBaseball/Book1.xlsx
workbook = openpyxl.load_workbook(path)
driver = webdriver.Firefox();
driver.get('')
# insert link to the scoreboard page of your fantasy league
try:
    # Automatic login
    WebDriverWait(driver, 1000).until(EC.presence_of_all_elements_located((By.XPATH, "(//iframe)")))
    driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
    time.sleep(2)

    driver.find_element_by_xpath("/html/body/div/div/div/section/section"
                                 "/form/section/div[1]/div/label/span[2]/input").send_keys('')
                                                                            # insert username/email in single quotes

    driver.find_element_by_xpath("/html/body/div/div/div/section/"
                                 "section/form/section/div[2]/div/label/span[2]/input").send_keys('')
                                                                            #insert password in single quotes

    driver.find_element_by_xpath("/html/body/div/div/div/section/section/form/section/div[3]/button").click()

    driver.switch_to.default_content()
    time.sleep(4)
finally:
    # a 'set' is a pair of managers and their stats for the week
    sets = driver.find_elements_by_xpath("//tbody[contains(@class, 'Table2__tbody')]")
    print("sets: " + str(len(sets)))

# get the select element, then the selected option
select = Select(driver.find_element_by_xpath("/html/body/div[1]/div[1]/div/div/div[5]"
                                             "/div[2]/div[1]/div/div/div[3]/div[1]/div/div[2]/select"))
option = select.first_selected_option

# matchupWeek will correspond to the row we need to write to in excel
matchupWeek = int(option.text.split(" ")[1])

print("Matchup Week: " + str(matchupWeek))
colCount = 2
worksheetCount = 1
for set in sets:
    # a row is one manager's week consisting of their nickname and their stats for each category
    rows = set.find_elements_by_xpath("./tr")
    for row in rows:
        cols = row.find_elements_by_xpath("./td")

        # Get opponent row to match with spreadsheet name
        opponent = rows[worksheetCount]
        opponentCols = opponent.find_elements_by_xpath("./td")
        sheet = workbook[opponentCols[0].text]

        # increment matchupWeek to account for header row in excel
        nextRow = matchupWeek + 1
        sheet.cell(row=nextRow, column=1, value=option.text)  # add week identifier(Ex. 'Matchup 1 (March X - March X)
        for col in cols:
            if colCount == 2:
                # the second column is the team nickname
                sheet.cell(row=nextRow, column=colCount, value=col.text)
            else:
                # float every numeric piece of data
                sheet.cell(row=nextRow, column=colCount, value=float(col.text))
            colCount += 1
        # decrement worksheetCount to switch to the second set of data. (The first set we get will now be the "opponent"
        worksheetCount -= 1
        colCount = 2
    worksheetCount = 1

# save workbook and close out the driver
driver.close()
workbook.save(path)
