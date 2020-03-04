import csv
import time
import io
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait # available since 2.4.0
from selenium.webdriver.support import expected_conditions as EC # available since 2.26.0
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

path = "C:\SeleniumPractice\Book1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

driver = webdriver.Firefox();
driver.get('https://fantasy.espn.com/baseball/league/standings?leagueId=192392&seasonId=2019')

# wait for page to load
time.sleep(15)
namesTable = driver.find_element_by_xpath('/html/body/div[1]/div[1]/div/div/div[5]/div[2]/div[2]/'
                                          'div/div/div[5]/section/table/tbody/tr/td[1]/div/table/tbody')
statsTable = driver.find_element_by_xpath('/html/body/div[1]/div[1]/div/div/div[5]/div[2]/div[2]/div/div/div[5]/section/'
                                          'table/tbody/tr/td[2]/div/div/div[2]/table/tbody/tr/td/div/table/tbody')
namesValues = namesTable.find_elements_by_xpath("./tr")
statsValues = statsTable.find_elements_by_xpath("./tr")

rowCount = 2
colCount = 1
for row in namesValues:
    for col in row.find_elements_by_xpath("./td"):
        sheet.cell(row=rowCount, column=colCount, value=col.text)
        colCount += 1
    rowCount += 1
    colCount = 1

rowCount = 2
colCount = 3
for row in statsValues:
    for col in row.find_elements_by_xpath("./td"):
        val = float(col.text)
        sheet.cell(row=rowCount, column=colCount, value=val)
        colCount += 1
    rowCount += 1
    colCount = 3

workbook.save(path)
driver.close()
